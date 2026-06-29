#!/usr/bin/env python3
"""Build the consolidated USLCI import trace + progress workbook.

Mirrors reports/bafu-import/build-bafu-trace-xlsx.py, adapted to the USLCI
(NREL openLCA JSON-LD) import workspace layout.

Sheets:
  1. 说明 Read me
  2. 导入进展 Summary       (coverage + flow/support/conversion totals)
  3. 关键决策 Key Decisions (the engineering decisions that closed the import)
  4. Process Trace          (every in-universe process, 1,358)
  5. Flow Trace             (every verified flow row, ~1,936)
  6. 转换映射 Conversion     (effective source-flow -> canonical reuse decisions)
  7. Support Identities     (contacts/sources/flowproperties/unitgroups verified)
"""
import json, os, re, glob
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Resolve paths relative to this script so the tool is committable and relocatable.
# Lives at <repo>/reports/uslci-import/build-uslci-trace-xlsx.py -> repo root is two dirs up.
# The run directory can be overridden with USLCI_RUN_DIR for a different import workspace.
HERE = os.path.dirname(os.path.abspath(__file__))
FOUNDRY = os.path.abspath(os.path.join(HERE, os.pardir, os.pardir))
RUN = os.environ.get(
    "USLCI_RUN_DIR",
    os.path.join(FOUNDRY, ".foundry/workspaces/uslci-full-import-20260612T093202Z"),
)
TIDAS_P = os.path.join(RUN, "conversion-v9/tidas/processes")
TIDAS_F = os.path.join(RUN, "conversion-v9/tidas/flows")
LEDGER = os.path.join(RUN, "batch-import-full-v2/import-ledger")
UNIVERSE_IDS = os.path.join(RUN, "universe-v1/universe-process-ids.txt")
READY_SCOPES = os.path.join(RUN, "library-resolution-v13/ready-scopes.jsonl")
RESOLUTION = os.path.join(RUN, "library-resolution-v20/exchange-reference-rewrites.jsonl")
OUT = os.path.join(HERE, "USLCI-导入trace与进展.xlsx")

# ---------------- helpers ----------------
def txt(x):
    if isinstance(x, list):
        for i in x:
            if isinstance(i, dict) and i.get("#text"):
                return i["#text"]
        return ""
    return x.get("#text", "") if isinstance(x, dict) else (str(x) if x is not None else "")

def loadl(p):
    out = []
    if os.path.exists(p):
        for ln in open(p):
            ln = ln.strip()
            if ln:
                try:
                    out.append(json.loads(ln))
                except Exception:
                    pass
    return out

def proc_name(pid):
    p = os.path.join(TIDAS_P, pid + ".json")
    if not os.path.exists(p):
        return ""
    try:
        d = json.load(open(p)); fl = d.get("processDataSet", d)
        return txt(fl["processInformation"]["dataSetInformation"]["name"]["baseName"])
    except Exception:
        return ""

def flow_name(fid):
    p = os.path.join(TIDAS_F, fid + ".json")
    if not os.path.exists(p):
        return ""
    try:
        d = json.load(open(p)); fl = d.get("flowDataSet", d)
        return txt(fl["flowInformation"]["dataSetInformation"]["name"]["baseName"])
    except Exception:
        return ""

# ---------------- load data ----------------
print("loading universe + ledgers ...")
universe_ids = [ln.strip() for ln in open(UNIVERSE_IDS)] if os.path.exists(UNIVERSE_IDS) else []
universe_ids = [u for u in universe_ids if u]
procs = loadl(os.path.join(LEDGER, "ok.processes.verified.jsonl"))
flows = loadl(os.path.join(LEDGER, "ok.flows.verified.jsonl"))
support = loadl(os.path.join(LEDGER, "verified-support-identities.jsonl"))

# process version from ready-scopes
pver = {}
for r in loadl(READY_SCOPES):
    pver[r.get("process_id")] = r.get("process_version", "")

# per-process verified meta from the ledger (dedup by process_id -> latest timestamp)
pmeta = {}  # process_id -> (version, timestamp)
for r in procs:
    pid = r.get("process_id")
    ts = (r.get("generated_at_utc") or "")[:19]
    if pid not in pmeta or ts > pmeta[pid][1]:
        pmeta[pid] = (r.get("dataset_version", "") or pver.get(pid, ""), ts)
verified_pids = set(pmeta)

# conversion decisions: dedup the per-exchange resolution to one effective row per source flow.
conv = {}   # source_flow_id -> row
for r in loadl(RESOLUTION):
    sid = r.get("source_flow_id")
    if not sid or sid in conv:
        continue
    conv[sid] = r
print(f"  universe={len(universe_ids)} verified_procs={len(verified_pids)} "
      f"flow_rows={len(flows)} support={len(support)} conv_source_flows={len(conv)}")

print(f"reading process names ({len(universe_ids)} tidas files) ...")
pname = {pid: proc_name(pid) for pid in universe_ids}

# ---------------- styles ----------------
HDR = PatternFill("solid", fgColor="1F4E78"); HDRF = Font(name="Arial", bold=True, color="FFFFFF", size=11)
VERI = PatternFill("solid", fgColor="E2EFDA"); PEND = PatternFill("solid", fgColor="FCE4D6")
DEC = PatternFill("solid", fgColor="FFF2CC"); SUBH = PatternFill("solid", fgColor="D9E1F2")
F = Font(name="Arial", size=10); FB = Font(name="Arial", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top"); TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D0D0D0"); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

def header(ws, cols, widths, row=1):
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ================= Sheet 1: 说明 =================
s = wb.active; s.title = "说明 Read me"
rows = [
 ("USLCI (NREL) 导入 — Trace 与进展总表", True, 15),
 ("", False, 10),
 ("账户 5c784552-09a5-43dc-b704-b96ed3239ecd (linanenv) | state_code=0 | 源 NREL USLCI openLCA JSON-LD (U.S. Federal LCA Commons)", False, 10),
 ("数据来源：.foundry/workspaces/uslci-full-import-20260612T093202Z（universe-v1 + batch-import-full-v2/import-ledger + library-resolution-v20 + conversion-v9/tidas 源）", False, 10),
 ("", False, 10),
 ("导入现状：universe 1,358 / verified 1,358（100%，全部入库已核验，0 残留）。", True, 11),
 ("", False, 10),
 ("各 sheet 说明：", True, 11),
 ("  • 导入进展 Summary —— 总体数字（universe/verified/flow/support/转换决策）与关键里程碑。", False, 10),
 ("  • 关键决策 Key Decisions —— 闭合本次导入的工程决策：land-use elementary 归类、canonical 源按 UUID 识别、17 个被 ST_200 遮盖的 flow bump 到 @01.00.001、142 个新建 flow。", False, 10),
 ("  • Process Trace —— 每个 in-universe process 一行(1,358)：id/版本/名称/状态/导入时间。", False, 10),
 ("  • Flow Trace —— 每条已核验 flow 一行(~1,936)：flow id/版本/所属 process/状态/时间。", False, 10),
 ("  • 转换映射 Conversion —— 每个源 flow 的有效解析决策：reuse(复用既有 canonical，含同 UUID 版本升级与跨 UUID 物理等价复用)。这是导入转换的核心 lineage。", False, 10),
 ("  • Support Identities —— 导入过程创建/核验的支撑数据集(~1,205)：contact/source/flow property/unit group 等。", False, 10),
 ("", False, 10),
 ("注：USLCI 与 ILCD canonical 参考库按 UUID 共享 elementary flow / 格式源(如 a97a0155 'ILCD format')，故这些一律复用既有 canonical 而非新建；少数与 ST_200 既有导入冲突的 flow 以 USLCI 为准 bump 到更高版本。", False, 9),
]
for i, (t, b, sz) in enumerate(rows, 1):
    c = s.cell(row=i, column=1, value=t); c.font = Font(name="Arial", bold=b, size=sz); c.alignment = WRAP
s.column_dimensions["A"].width = 130

# ================= Sheet 2: 导入进展 Summary =================
s = wb.create_sheet("导入进展 Summary")
ver = sum(1 for pid in universe_ids if pid in verified_pids)
pend = len(universe_ids) - ver
kv = [
 ("指标", "值"),
 ("Process universe (总过程数)", len(universe_ids)),
 ("Verified (已入库已核验)", ver),
 ("Pending (待处理残留)", pend),
 ("Coverage %", f"{ver/len(universe_ids)*100:.3f}%" if universe_ids else "n/a"),
 ("Flow rows verified (已核验 flow 行)", len(flows)),
 ("Support identities verified (支撑数据集)", len(support)),
 ("Conversion source flows (有效源 flow 解析决策)", len(conv)),
]
for r in kv:
    s.append(list(r))
for c in range(1, 3):
    s.cell(row=1, column=c).fill = HDR; s.cell(row=1, column=c).font = HDRF
for r in range(2, len(kv) + 1):
    s.cell(row=r, column=1).font = FB
s.column_dimensions["A"].width = 46; s.column_dimensions["B"].width = 22
# conversion decision-type distribution
s.append([]); s.append(["转换决策类型分布 (effective source-flow resolution)", ""])
s.cell(row=s.max_row, column=1).font = FB; s.cell(row=s.max_row, column=1).fill = SUBH
s.append(["决策类型", "源 flow 数"])
for c in range(1, 3):
    s.cell(row=s.max_row, column=c).fill = HDR; s.cell(row=s.max_row, column=c).font = HDRF
def conv_kind(r):
    same = r.get("source_flow_id") == r.get("canonical_flow_id")
    vbump = r.get("source_flow_version") != r.get("canonical_flow_version")
    if same and vbump:
        return "reuse 同UUID版本升级 (canonical 高版本)"
    if same:
        return "reuse 同UUID同版本"
    return "reuse 跨UUID物理等价 (canonical 不同 flow)"
ck = Counter(conv_kind(r) for r in conv.values())
for k, v in ck.most_common():
    s.append([k, v])

# ================= Sheet 3: 关键决策 Key Decisions =================
s = wb.create_sheet("关键决策 Key Decisions")
s.append(["类别", "对象 / 范围", "决策 / 结论", "说明"])
for c in range(1, 5):
    s.cell(row=1, column=c).fill = HDR; s.cell(row=1, column=c).font = HDRF
s.row_dimensions[1].height = 28; s.freeze_panes = "A2"
kd = [
 ("① 单位归一化", "全部 flow property / unit group", "APPLIED ✓",
  "tidas-tools 单位归一化闭环（Phase 1+2），source unit -> canonical reference unit。"),
 ("② land-use elementary 归类", "117 个 land-transformation flow", "APPLIED ✓",
  "USLCI 源把 land-transformation flow 标成 PRODUCT_FLOW/Ecosystem Services；tidas-tools 按名识别(Land use 3.2/3.1) promote 成 elementary，改走 elementary reuse-matching。"),
 ("③ canonical 源按 UUID 识别", "a97a0155 (ILCD format) / d92a1a12 (ILCD 合规)", "APPLIED ✓ (本会话代码修复)",
  "源把 'ILCD format' 标成 'Publications and communications' 分类并作 referenceToDataSource，使 sourceSemanticKind 误判为 placeholder 而新建@00.00.001，触发 source 身份门并会 version_outdated。修复：按已知 canonical 源 UUID 识别 -> 复用 @03.00.003，不新建。(scripts/lib/source-semantics.mjs)"),
 ("④ 物理等价 elementary 复用", "decisions-v6: 3,380 reuse / 539 mint", "APPLIED ✓",
  "AI-first 物理等价匹配复用既有 canonical elementary flow；184 个 synonym 救回。"),
 ("⑤ 142 个新建 flow (flow-first mint)", "128 elementary + 8 product + 6 waste", "MINTED ✓ @00.00.001",
  "库内无质量等价 canonical 的源 flow，经对抗复核确认确为新增后批量 flow publish-version 新建（清理 import-trace 后过 publish 门）。"),
 ("⑥ 11 个 mega-scope 闭合", "1,191–2,429 exchanges/scope", "VERIFIED ✓",
  "最重的 11 个 scope 因复用闭包与源身份门曾被 defer；经 ③ 修复 + ⑦ bump 后全部入库。"),
 ("⑦ ST_200 版本冲突 -> bump", "17 个 flow (ST_200 9c4751b6 持 @01.00.000)", "BUMPED ✓ -> @01.00.001",
  "这 17 个 flow 在用户自有 ST_200 账户已发布 @01.00.000，遮盖我们导入的 @00.00.001 -> version_outdated。平台只能删草稿(state_code=0)无法删已发布行，故按'USLCI 为准'把我们的副本 bump 到 @01.00.001(全局最新)，并把引用重指到 @01.00.001。"),
]
for row in kd:
    s.append(row)
for r in range(2, s.max_row + 1):
    for c in range(1, 5):
        cell = s.cell(row=r, column=c); cell.font = F; cell.alignment = WRAP; cell.border = BORD
    st = s.cell(row=r, column=3).value or ""
    s.cell(row=r, column=3).fill = VERI if ("✓" in st) else DEC
for i, w in enumerate([26, 40, 30, 90], 1):
    s.column_dimensions[get_column_letter(i)].width = w

# ================= Sheet 4: Process Trace =================
s = wb.create_sheet("Process Trace")
header(s, ["process_id", "version", "名称 baseName", "状态", "导入时间 UTC"],
       [40, 12, 64, 12, 22])
for pid in universe_ids:
    ver_v, ts = pmeta.get(pid, (pver.get(pid, ""), ""))
    st = "verified" if pid in verified_pids else "pending"
    s.append([pid, ver_v, pname.get(pid, ""), st, ts])
s.auto_filter.ref = f"A1:E{s.max_row}"
for r in range(2, s.max_row + 1):
    stv = s.cell(row=r, column=4).value
    s.cell(row=r, column=4).fill = VERI if stv == "verified" else PEND

# ================= Sheet 5: Flow Trace =================
s = wb.create_sheet("Flow Trace")
header(s, ["flow_id", "version", "所属 process_id", "状态", "导入时间 UTC"],
       [40, 12, 40, 12, 22])
for r in flows:
    s.append([r.get("dataset_id", ""), r.get("dataset_version", ""), r.get("process_id", ""),
              r.get("status", ""), (r.get("generated_at_utc") or "")[:19]])
s.auto_filter.ref = f"A1:E{s.max_row}"

# ================= Sheet 6: 转换映射 Conversion =================
s = wb.create_sheet("转换映射 Conversion")
header(s, ["源 flow id (USLCI)", "源 flow 名称", "源版本", "决策类型", "→ canonical id",
           "canonical 名称", "canonical 版本"],
       [40, 34, 12, 34, 40, 34, 12])
print(f"reading {len(conv)} source-flow names for conversion sheet ...")
for sid, r in sorted(conv.items()):
    s.append([sid, flow_name(sid), r.get("source_flow_version", ""), conv_kind(r),
              r.get("canonical_flow_id", ""),
              r.get("canonical_short_description", "") or flow_name(r.get("canonical_flow_id", "")),
              r.get("canonical_flow_version", "")])
s.auto_filter.ref = f"A1:G{s.max_row}"

# ================= Sheet 7: Support Identities =================
s = wb.create_sheet("Support Identities")
header(s, ["identity_key", "dataset_type", "dataset_id", "version", "状态", "来源 source"],
       [50, 16, 40, 12, 12, 30])
for r in support:
    s.append([r.get("identity_key", ""), r.get("dataset_type", ""), r.get("dataset_id", ""),
              r.get("dataset_version", ""), r.get("status", ""), r.get("source", "")])
s.auto_filter.ref = f"A1:F{s.max_row}"

# ---------------- save ----------------
print("saving workbook ...")
wb.save(OUT)
print("SAVED:", OUT)
print("SHEET ROWS:", {ws.title: ws.max_row for ws in wb.worksheets})
print("VERIFY_COUNTS", json.dumps({
    "universe": len(universe_ids), "verified": ver, "pending": pend,
    "flows": len(flows), "support": len(support), "conversion_source_flows": len(conv),
}))
